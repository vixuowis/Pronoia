"""build_assessment_panels_xlsx.py — Pronoia-RLVR §5.2 评估面板 Excel 生成。

把 eval_rlvr_vs_baseline.py 输出的 JSON 报告转成多 sheet xlsx：
  Sheet1 OVERVIEW        ：RLVR vs 四基线，primary/avg_all/双窗 三主指标 + Wilson CI
  Sheet2 VOLUME_3BUCKET  ：HIGH/NORMAL/LOW 三桶 ACC × 各模型
  Sheet3 RER_5H          ：5 个 horizon（t3/t7/t15/t30/t60）× RER 均值 + 同号率 + ACC_vs_RER
  Sheet4 SCENE_BUCKET    ：Market × EventType × vol_regime 3D 小格 ACC（heatmap 数值）
  Sheet5 MOE_HEALTH      ：Router 熵 + active experts 分布
  Sheet6 COMBINED_GATES  ：8 条组合门禁 PASS/FAIL（§5.3）

依赖：openpyxl（pip install openpyxl，轻量）。没有就降级输出 JSON。
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path


def _check_openpyxl() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except Exception:
        return False


def _8_combined_gates(report: dict) -> dict[str, dict]:
    """§5.3 八条组合门禁（RLVR 才能上线的硬门槛）。"""
    rlvr = report.get("RLVR") or {}
    b0   = report.get("B0_majority") or {}
    b2   = report.get("B2_oracle_primary") or {}

    def _acc(x): return (x or {}).get("value", 0.0)

    def _wilson_sig(rlvr_acc, base_acc, rlvr_n, base_n):
        """Wilson CI 显著优于基线（RLVR 下界 > 基线 上界 的简化版本，用 n 近似）。"""
        import math
        if rlvr_n == 0: return False
        z = 1.96
        def _ci(k, n):
            if n == 0: return (0,0)
            p = k/n; kk = int(p*n)
            denom = 1 + z*z/n
            c = (p + z*z/(2*n)) / denom
            m = z * math.sqrt(p*(1-p)/n + z*z/(4*n*n)) / denom
            return c-m, c+m
        rlvr_lo, _ = _ci(int(rlvr_acc * rlvr_n), rlvr_n)
        _, base_hi = _ci(int(base_acc * base_n), base_n)
        return rlvr_lo > base_hi

    n = rlvr.get("n_matched_pairs", 0)
    rlvr_primary = _acc(rlvr.get("primary_strict_ACC"))
    b0_primary   = _acc(b0.get("primary_strict_ACC"))

    gates = {}
    # G1: 定向 primary ACC 显著优于 B0 majority（Wilson）
    g1 = _wilson_sig(rlvr_primary, b0_primary, n, n) if (b0_primary > 0 and n > 0) else False
    gates["G1_primary_gt_B0_majority"] = {"pass": g1, "RLVR": rlvr_primary, "B0": b0_primary, "note": "Wilson 95% CI 显著高于 baseline"}

    # G2: primary ACC ≥ B2 oracle × 75%（理论上界 75% 以上 = 接近上限）
    b2_primary = _acc(b2.get("primary_strict_ACC"))
    g2 = (b2_primary > 0) and (rlvr_primary >= b2_primary * 0.75)
    gates["G2_primary_ge_75pct_oracle"] = {"pass": g2, "RLVR": rlvr_primary, "B2_oracle": b2_primary, "ratio": rlvr_primary/max(1e-9,b2_primary)}

    # G3: avg_all ACC ≥ 旧 baseline（若提供 baseline Tier1 则比，否则比 B0 + 3pp）
    rlvr_avg = _acc(rlvr.get("avg_all_strict_ACC"))
    t1 = report.get("B3_Tier1") or report.get("Tier1")
    tier1_avg = _acc((t1 or {}).get("avg_all_strict_ACC"))
    if tier1_avg > 0:
        g3 = rlvr_avg >= tier1_avg + 0.03
        gates["G3_avgall_ge_Tier1_3pp"] = {"pass": g3, "RLVR": rlvr_avg, "Tier1": tier1_avg}
    else:
        g3 = rlvr_avg >= b0_primary + 0.03
        gates["G3_avgall_ge_B0_3pp"] = {"pass": g3, "RLVR": rlvr_avg, "B0": b0_primary}

    # G4: 双窗一致率 ≥ 60%
    dw = (rlvr.get("double_window_consistency") or {}).get("value", 0.0)
    g4 = dw >= 0.60
    gates["G4_double_window_ge_60pct"] = {"pass": g4, "value": dw}

    # G5: vol_regime HIGH 桶 ACC ≥ NORMAL 桶 ACC × 0.9（不能因为量价高波动而掉很多）
    bk = rlvr.get("volume_3bucket_ACC") or {}
    acc_high   = (bk.get("HIGH") or {}).get("acc", 0.0)
    acc_normal = (bk.get("NORMAL") or {}).get("acc", 0.0)
    g5 = (acc_normal > 0) and (acc_high >= acc_normal * 0.90)
    gates["G5_HIGH_bucket_ge_90pct_NORMAL"] = {"pass": g5, "HIGH_acc": acc_high, "NORMAL_acc": acc_normal}

    # G6: RER↔CAR t7 同号率 ≥ 70%（alpha 纯度）
    rp = rlvr.get("rer_panel") or {}
    agree_t7 = ((rp.get("t7") or {}).get("RER_CAR_agree_ratio", 0.0))
    g6 = agree_t7 >= 0.70
    gates["G6_RER_CAR_agree_t7_ge_70pct"] = {"pass": g6, "value": agree_t7}

    # G7: 长短 horizon 反转：t3 vs t60 rer_car_agree 差 < 10pp
    agree_t3  = ((rp.get("t3")  or {}).get("RER_CAR_agree_ratio", 0.0))
    agree_t60 = ((rp.get("t60") or {}).get("RER_CAR_agree_ratio", 0.0))
    g7 = abs(agree_t3 - agree_t60) <= 0.10
    gates["G7_RER_agree_t3_t60_gap_le_10pp"] = {"pass": g7, "t3": agree_t3, "t60": agree_t60, "gap_abs": abs(agree_t3-agree_t60)}

    # G8: MoE router_entropy ∈ [0.4*ln6, 0.85*ln6]（≈ [0.716, 1.521]）
    import math
    H_max = math.log(6); lo = 0.4*H_max; hi = 0.85*H_max
    ent = float((rlvr.get("moe_health") or {}).get("router_entropy_mean", 0.0))
    g8 = lo <= ent <= hi
    gates["G8_MoE_router_entropy_in_range"] = {"pass": g8, "value": ent, "expected_range": [round(lo,3), round(hi,3)]}

    gates["_ALL_GATES"] = {"pass": all(g["pass"] for k, g in gates.items() if not k.startswith("_")),
                           "n_passed": sum(1 for k, g in gates.items() if not k.startswith("_") and g["pass"]),
                           "n_total":  len([k for k in gates if not k.startswith("_")])}
    return gates


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--report-json", required=True)
    ap.add_argument("--out-xlsx", required=True)
    args = ap.parse_args()

    report = json.load(open(args.report_json, encoding="utf-8"))
    gates  = _8_combined_gates(report)
    has_xl = _check_openpyxl()

    out_path = Path(args.out_xlsx)

    if not has_xl:
        # 降级：输出 {out-xlsx}.gates.json
        alt = out_path.with_suffix(".gates.json")
        with open(alt, "w", encoding="utf-8") as f:
            json.dump({"combined_gates": gates, "report_keys": list(report.keys())}, f, ensure_ascii=False, indent=2)
        print(f"[WARN] openpyxl 未安装，跳过 xlsx 生成；门禁 JSON → {alt}")
        print(f"       安装：pip install openpyxl，再重新运行即可生成 xlsx 面板。")
        return

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    pass_fill   = PatternFill("solid", fgColor="C6EFCE")
    fail_fill   = PatternFill("solid", fgColor="FFC7CE")
    center = Alignment(horizontal="center", vertical="center")

    def _write_header(ws, cols, row=1):
        for j, c in enumerate(cols, 1):
            cell = ws.cell(row=row, column=j, value=c)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = center

    # ========= Sheet1 OVERVIEW =========
    ws = wb.active; ws.title = "OVERVIEW"
    _write_header(ws, ["Model", "primary ACC (%)", "primary CI",
                        "avg_all ACC (%)", "avg_all CI",
                        "double-window consistency (%)", "n_pairs"])
    row = 2
    for name, rr in report.items():
        if not isinstance(rr, dict) or "n_matched_pairs" not in rr: continue
        p1 = rr["primary_strict_ACC"]["value"]*100
        ci1 = f"[{rr['primary_strict_ACC']['CI'][0]*100:.1f}, {rr['primary_strict_ACC']['CI'][1]*100:.1f}]"
        p2 = rr["avg_all_strict_ACC"]["value"]*100
        ci2 = f"[{rr['avg_all_strict_ACC']['CI'][0]*100:.1f}, {rr['avg_all_strict_ACC']['CI'][1]*100:.1f}]"
        dw = rr["double_window_consistency"]["value"]*100
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=round(p1,2)).alignment = center
        ws.cell(row=row, column=3, value=ci1)
        ws.cell(row=row, column=4, value=round(p2,2)).alignment = center
        ws.cell(row=row, column=5, value=ci2)
        ws.cell(row=row, column=6, value=round(dw,2)).alignment = center
        ws.cell(row=row, column=7, value=rr["n_matched_pairs"]).alignment = center
        row += 1
    for col in range(1, 8): ws.column_dimensions[get_column_letter(col)].width = 28

    # ========= Sheet2 VOLUME_3BUCKET =========
    ws2 = wb.create_sheet("VOLUME_3BUCKET")
    _write_header(ws2, ["Model", "HIGH_acc", "HIGH_n", "NORMAL_acc", "NORMAL_n", "LOW_acc", "LOW_n"])
    row = 2
    for name, rr in report.items():
        if not isinstance(rr, dict) or "volume_3bucket_ACC" not in rr: continue
        b = rr["volume_3bucket_ACC"] or {}
        vals = [name]
        for rg in ("HIGH", "NORMAL", "LOW"):
            x = b.get(rg) or {}
            vals += [round(x.get("acc",0)*100,2), x.get("n",0)]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(row=row, column=j, value=v); c.alignment = center
        row += 1
    for col in range(1, 8): ws2.column_dimensions[get_column_letter(col)].width = 16

    # ========= Sheet3 RER_5H =========
    ws3 = wb.create_sheet("RER_5H")
    headers = ["Model", "Horizon", "n_valid", "RER_mean(%)", "CAR_mean(%)",
               "RER_pos_ratio", "RER_CAR_agree", "ACC_vs_RER"]
    _write_header(ws3, headers)
    row = 2
    for name, rr in report.items():
        if not isinstance(rr, dict) or "rer_panel" not in rr: continue
        rp = rr["rer_panel"] or {}
        for h in ["t3","t7","t15","t30","t60"]:
            d = rp.get(h) or {}
            vals = [name, h, d.get("n_valid",0), d.get("RER_mean_pct",0), d.get("CAR_mean_pct",0),
                    d.get("RER_positive_ratio",0), d.get("RER_CAR_agree_ratio",0), d.get("ACC_vs_RER_direction",0)]
            for j, v in enumerate(vals, 1):
                c = ws3.cell(row=row, column=j, value=v); c.alignment = center
            row += 1
    for col in range(1, len(headers)+1): ws3.column_dimensions[get_column_letter(col)].width = 18

    # ========= Sheet4 SCENE_BUCKET =========
    ws4 = wb.create_sheet("SCENE_BUCKET")
    _write_header(ws4, ["Market|Event|Regime", "RLVR_n", "RLVR_acc", "CI"])
    row = 2
    rlvr = report.get("RLVR") or {}
    for k, v in (rlvr.get("scene_bucket_ACC") or {}).items():
        ci = v.get("CI") or []
        ws4.cell(row=row, column=1, value=k)
        ws4.cell(row=row, column=2, value=v.get("n",0)).alignment = center
        ws4.cell(row=row, column=3, value=v.get("acc",0)).alignment = center
        ws4.cell(row=row, column=4, value=f"[{ci[0] if ci else '?'}, {ci[1] if ci else '?'}]").alignment = center
        row += 1
    for col in range(1, 5): ws4.column_dimensions[get_column_letter(col)].width = 30

    # ========= Sheet5 MOE_HEALTH =========
    ws5 = wb.create_sheet("MOE_HEALTH")
    _write_header(ws5, ["Metric", "Value"])
    row = 2
    mh = (report.get("RLVR") or {}).get("moe_health") or {}
    for k, v in mh.items():
        if isinstance(v, dict): v = json.dumps(v, ensure_ascii=False)
        ws5.cell(row=row, column=1, value=k)
        ws5.cell(row=row, column=2, value=str(v)).alignment = center
        row += 1
    ws5.column_dimensions["A"].width = 40; ws5.column_dimensions["B"].width = 60

    # ========= Sheet6 COMBINED_GATES =========
    ws6 = wb.create_sheet("COMBINED_GATES")
    _write_header(ws6, ["Gate", "Pass?", "Detail"])
    row = 2
    for name, g in gates.items():
        c_pass = ws6.cell(row=row, column=1, value=name)
        val = g.get("pass")
        c1 = ws6.cell(row=row, column=2, value="✅ PASS" if val else ("❌ FAIL" if isinstance(val,bool) else ""))
        c1.alignment = center
        if isinstance(val, bool):
            c1.fill = pass_fill if val else fail_fill
        ws6.cell(row=row, column=3, value=json.dumps({k:v for k,v in g.items() if k != "pass"}, ensure_ascii=False))
        row += 1
    ws6.column_dimensions["A"].width = 44
    ws6.column_dimensions["B"].width = 14
    ws6.column_dimensions["C"].width = 100

    wb.save(str(out_path))
    print(f"[DONE] 评估面板 → {out_path}（6 sheets）")


if __name__ == "__main__":
    main()
